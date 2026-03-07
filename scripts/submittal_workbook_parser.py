#!/usr/bin/env python3
"""
Parse Submittal Manager.xlsm into normalized agency -> permit -> required-item structures.

Usage:
  python scripts/submittal_workbook_parser.py --workbook "z:\\Shared\\Standards\\Submittal Manager\\Submittal Manager.xlsm" --out-json scripts/submittal_workbook_dry_run.json --out-md scripts/submittal_workbook_dry_run.md
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


SPECIAL_MARKERS = (
    "OPEN",
    "ITEM",
    "REQUIRED UPLOADS",
    "TOTAL DOCUMENTS",
    "IF APPLICABLE",
)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
    else:
        text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def slugify(text: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", text.strip().lower()).strip("_")
    return slug or "item"


def is_special_marker(text: str) -> bool:
    if not text:
        return True
    upper = text.upper().rstrip(":")
    if upper in SPECIAL_MARKERS:
        return True
    if re.fullmatch(r"PHASE\s*\d+", upper):
        return True
    if upper.startswith("PHASE "):
        return True
    if upper.startswith("TOTAL DOCUMENT"):
        return True
    return False


def guess_item_type(text: str) -> str:
    upper = text.upper()
    if "APPLICATION" in upper or "CHECKLIST" in upper:
        return "application"
    if "PLAN" in upper or "CALC" in upper or "REPORT" in upper:
        return "plan"
    return "document"


@dataclass
class ItemRecord:
    agency_code: str
    agency_name: str
    permit_name: str
    permit_code: str
    permit_column: int
    row_number: int
    item_name: str
    item_code: str
    item_type_guess: str
    flag_type: Optional[str] = None
    notes: Optional[str] = None


def detect_header_row(ws: Any, max_rows: int = 12, max_cols: int = 40) -> Tuple[int, List[Tuple[int, str]]]:
    first_single: Optional[Tuple[int, List[Tuple[int, str]]]] = None
    for r in range(1, min(max_rows, ws.max_row) + 1):
        candidates: List[Tuple[int, str]] = []
        for c in range(1, min(max_cols, ws.max_column) + 1):
            text = normalize_text(ws.cell(r, c).value)
            if text and not is_special_marker(text):
                candidates.append((c, text))

        if not candidates:
            continue

        # Header rows in this workbook appear near the top and usually have
        # multiple permit names spread across columns.
        if len(candidates) >= 2:
            return r, candidates
        if first_single is None:
            first_single = (r, candidates)

    if first_single is not None:
        return first_single
    return 1, []


def extract_item_from_block(ws: Any, row: int, start_col: int, end_col: int) -> str:
    for c in range(start_col, end_col + 1):
        text = normalize_text(ws.cell(row, c).value)
        if text:
            return text
    return ""


def parse_sheet(ws: Any) -> Dict[str, Any]:
    header_row, headers = detect_header_row(ws)
    if not headers:
        return {
            "agency_code": ws.title,
            "agency_name": ws.title,
            "header_row": header_row,
            "permits": [],
            "items": [],
            "flagged_rows": [],
            "notes": ["No permit headers detected"],
        }

    sorted_headers = sorted(headers, key=lambda x: x[0])
    permits: List[Dict[str, Any]] = []
    items: List[ItemRecord] = []
    flagged: List[ItemRecord] = []

    for i, (col, permit_name) in enumerate(sorted_headers):
        next_col = sorted_headers[i + 1][0] - 1 if i + 1 < len(sorted_headers) else min(ws.max_column, col + 5)
        permit_code = slugify(permit_name).upper()
        permits.append(
            {
                "permit_name": permit_name,
                "permit_code": permit_code,
                "start_col": col,
                "end_col": next_col,
            }
        )

        for r in range(header_row + 1, ws.max_row + 1):
            candidate = extract_item_from_block(ws, r, col, next_col)
            if not candidate:
                continue

            if normalize_text(candidate).upper() == normalize_text(permit_name).upper():
                flagged.append(
                    ItemRecord(
                        agency_code=ws.title,
                        agency_name=ws.title,
                        permit_name=permit_name,
                        permit_code=permit_code,
                        permit_column=col,
                        row_number=r,
                        item_name=candidate,
                        item_code=slugify(candidate).upper(),
                        item_type_guess="document",
                        flag_type="repeated_header",
                        notes="Value matches permit name",
                    )
                )
                continue

            record = ItemRecord(
                agency_code=ws.title,
                agency_name=ws.title,
                permit_name=permit_name,
                permit_code=permit_code,
                permit_column=col,
                row_number=r,
                item_name=candidate,
                item_code=slugify(candidate).upper(),
                item_type_guess=guess_item_type(candidate),
            )
            if is_special_marker(candidate):
                record.flag_type = "special_marker"
                record.notes = "Needs explicit decision"
                flagged.append(record)
            else:
                items.append(record)

    return {
        "agency_code": ws.title,
        "agency_name": ws.title,
        "header_row": header_row,
        "permits": permits,
        "items": [asdict(x) for x in items],
        "flagged_rows": [asdict(x) for x in flagged],
        "notes": [],
    }


def build_markdown_report(parsed: Dict[str, Any]) -> str:
    lines: List[str] = ["# Submittal Workbook Dry-Run Report", ""]
    lines.append(f"- Workbook: `{parsed['workbook_path']}`")
    lines.append(f"- Agencies parsed: `{len(parsed['agencies'])}`")
    lines.append(f"- Actionable items: `{parsed['totals']['actionable_items']}`")
    lines.append(f"- Flagged rows: `{parsed['totals']['flagged_rows']}`")
    lines.append("")

    for agency in parsed["agencies"]:
        lines.append(f"## {agency['agency_code']}")
        lines.append(f"- Detected header row: `{agency['header_row']}`")
        lines.append(f"- Permits: `{len(agency['permits'])}`")
        lines.append(f"- Actionable items: `{len(agency['items'])}`")
        lines.append(f"- Flagged rows: `{len(agency['flagged_rows'])}`")
        lines.append("")

        permit_counts: Dict[str, int] = {}
        for item in agency["items"]:
            permit_counts[item["permit_name"]] = permit_counts.get(item["permit_name"], 0) + 1
        for permit in agency["permits"]:
            count = permit_counts.get(permit["permit_name"], 0)
            lines.append(f"- `{permit['permit_name']}`: `{count}` actionable items")

        if agency["flagged_rows"]:
            lines.append("")
            lines.append("### Flagged Rows")
            for row in agency["flagged_rows"][:40]:
                lines.append(
                    f"- row `{row['row_number']}` / permit `{row['permit_name']}`: "
                    f"`{row['item_name']}` (`{row.get('flag_type')}`)"
                )
            if len(agency["flagged_rows"]) > 40:
                lines.append(f"- ... and `{len(agency['flagged_rows']) - 40}` more")
        lines.append("")

    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, help="Path to Submittal Manager workbook (.xlsm)")
    parser.add_argument("--out-json", required=True, help="Output JSON path")
    parser.add_argument("--out-md", required=True, help="Output markdown report path")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    wb = openpyxl.load_workbook(str(workbook_path), data_only=True, read_only=True, keep_vba=True)

    agencies = [parse_sheet(ws) for ws in wb.worksheets]
    result = {
        "workbook_path": str(workbook_path),
        "agencies": agencies,
        "totals": {
            "actionable_items": sum(len(a["items"]) for a in agencies),
            "flagged_rows": sum(len(a["flagged_rows"]) for a in agencies),
        },
    }

    out_json = Path(args.out_json)
    out_json.write_text(json.dumps(result, indent=2), encoding="utf-8")

    out_md = Path(args.out_md)
    out_md.write_text(build_markdown_report(result), encoding="utf-8")

    print(
        json.dumps(
            {
                "agencies": len(agencies),
                "actionable_items": result["totals"]["actionable_items"],
                "flagged_rows": result["totals"]["flagged_rows"],
                "out_json": str(out_json),
                "out_md": str(out_md),
            }
        )
    )


if __name__ == "__main__":
    main()
